from openpyxl import load_workbook
import pandas as pd

print("####################\n# Redirect 301     #\n# by Markus Jahnel #\n####################")

# User input
filename = input("Enter a filename. The file must be located in the files folder: ")
filename = "files/" + filename

data_sheet = input("Enter the name of the Excel-Sheet e.g. Tabelle1: ")

url = input("Enter a valid URL e.g. https://www.domain.de: ")


# Open and parse Excel file
xl = pd.ExcelFile(filename)
df = xl.parse(data_sheet)

# Sort Excel-Sheet
df = df.sort_values(by=["Old URL"], ascending=False)
writer = pd.ExcelWriter('files/redirects_sort.xlsx')
df.to_excel(writer, sheet_name=data_sheet, columns=["Old URL", "New URL"], index=False)
writer.save()

# Search and replace URL
workbook = load_workbook(filename="files/redirects_sort.xlsx")

ws = workbook[data_sheet]

i = 0
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+0):
        s = ws.cell(r, c).value
        if s != None and url in s:
            ws.cell(r, c).value = s.replace(url, "Redirect 301 ")
            print("row {} col {} : {}".format(r, c, s))
            i += 1
        else:
            print("URL not found in Excel-Sheet")

workbook.save('files/redirects_final.xlsx')
